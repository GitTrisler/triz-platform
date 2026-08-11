"""
TRIZ Project Hub — deliverable package builder.

Copies the newest revision of every (optionally filtered) document into a
dated package folder and writes a transmittal sheet. Deliberately minimal —
the full plot/publish pipeline lives in the TRIZ Deliverable Publisher module;
this covers the "hand me the current set" case straight from the index.
"""

from __future__ import annotations

import shutil
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def build_package(db, project_root: str, out_dir: str,
                  doc_class: str | None = None, ext: str | None = ".pdf") -> dict:
    root = Path(project_root)
    stamp = time.strftime("%Y-%m-%d")
    pkg = Path(out_dir) / f"Package_{stamp}"
    pkg.mkdir(parents=True, exist_ok=True)

    docs = db.latest_documents()
    copied, skipped = [], []
    for d in docs:
        if doc_class and (d["doc_class"] or "") != doc_class:
            continue
        if ext and not d["path"].lower().endswith(ext):
            skipped.append(d)
            continue
        src = root / d["path"]
        if not src.exists():
            skipped.append(d)
            continue
        dest = pkg / src.name
        shutil.copy2(src, dest)
        copied.append(d)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transmittal"
    ws.append([f"Deliverable Package — {stamp}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Document No", "Rev", "Title", "File"])
    for c in ws[3]:
        c.fill = PatternFill("solid", fgColor="1F2430")
        c.font = Font(color="FFFFFF", bold=True)
    for d in copied:
        ws.append([d["doc_number"], d["revision"] or "", d["title"] or "", d["name"]])
    for i, w in enumerate([18, 6, 45, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(str(pkg / f"Transmittal_{stamp}.xlsx"))

    return {"folder": str(pkg), "copied": len(copied), "skipped": len(skipped)}

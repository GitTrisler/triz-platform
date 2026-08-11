"""
TRIZ Project Hub — Excel extractor (openpyxl).

Three behaviors per sheet:
  1. Register detection  — a header row containing drawing/doc + rev columns
                           means the sheet is a drawing register: rows become
                           document records.
  2. Model-export detection — a header row with tag + class/description/size
                           (Plant 3D / Navisworks exports) marks every tag as
                           existing in the model (objects.in_model = 1). This
                           is what powers the "P&ID tag missing from model"
                           check.
  3. Tag harvesting      — every string cell is run through the tag patterns;
                           hits become occurrences located as Sheet!Cell.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .patterns import TagMatcher, context_snippet
from .util import winsafe

MAX_ROWS_PER_SHEET = 5000
MAX_SCAN_COLS = 40

_REGISTER_DOC_HEADERS = ("drawing number", "drawing no", "doc number", "doc no",
                         "document number", "document no", "dwg no", "dwg number")
_REGISTER_REV_HEADERS = ("rev", "revision", "rev no")
_REGISTER_TITLE_HEADERS = ("title", "description", "drawing title", "sheet title")
_MODEL_TAG_HEADERS = ("tag", "tag number", "equipment tag", "item tag")
_MODEL_HINT_HEADERS = ("class", "class name", "description", "size", "line number",
                       "spec", "service", "part family")


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _find_header(rows_iter):
    """Look in the first 10 rows for a header. Returns (header_row_index, headers) or (None, None)."""
    cached = []
    for i, row in enumerate(rows_iter):
        cached.append(row)
        if i >= 9:
            break
    for i, row in enumerate(cached):
        headers = [_norm(c) for c in row[:MAX_SCAN_COLS]]
        if any(h in _REGISTER_DOC_HEADERS for h in headers) or any(h in _MODEL_TAG_HEADERS for h in headers):
            return i, headers, cached
    return None, None, cached


def extract_xlsx(abs_path: Path, file_id: int, db, matcher: TagMatcher, log=None):
    wb = load_workbook(winsafe(abs_path), read_only=True, data_only=True)
    stats = {"register_rows": 0, "model_tags": 0, "tags": 0, "is_register": False,
             "is_model_export": False}
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(max_row=MAX_ROWS_PER_SHEET, max_col=MAX_SCAN_COLS,
                                values_only=True)
            header_idx, headers, cached = _find_header(rows)

            doc_col = rev_col = title_col = tag_col = None
            mode = None
            if headers:
                for ci, h in enumerate(headers):
                    if h in _REGISTER_DOC_HEADERS and doc_col is None:
                        doc_col = ci
                    if h in _REGISTER_REV_HEADERS and rev_col is None:
                        rev_col = ci
                    if h in _REGISTER_TITLE_HEADERS and title_col is None:
                        title_col = ci
                    if h in _MODEL_TAG_HEADERS and tag_col is None:
                        tag_col = ci
                if doc_col is not None and rev_col is not None:
                    mode = "register"
                    stats["is_register"] = True
                elif tag_col is not None and any(h in _MODEL_HINT_HEADERS for h in headers):
                    mode = "model_export"
                    stats["is_model_export"] = True

            def handle_row(row_vals, row_no):
                # Structured ingestion
                if mode == "register" and row_vals[doc_col] not in (None, ""):
                    doc = str(row_vals[doc_col]).strip().upper()
                    rev = row_vals[rev_col]
                    rev = str(rev).strip().upper() if rev not in (None, "") else None
                    title = None
                    if title_col is not None and row_vals[title_col] not in (None, ""):
                        title = str(row_vals[title_col]).strip()
                    db.add_document(file_id, doc, title, rev, "excel_register")
                    stats["register_rows"] += 1
                if mode == "model_export" and row_vals[tag_col] not in (None, ""):
                    tag_text = str(row_vals[tag_col]).strip().upper()
                    found = matcher.find_unique(tag_text)
                    if found:
                        for t, tm in found.items():
                            db.mark_in_model(t, tm.object_type)
                    else:
                        db.mark_in_model(tag_text, "equipment")
                    stats["model_tags"] += 1
                # Tag harvesting across every string cell
                for ci, val in enumerate(row_vals):
                    if not isinstance(val, str) or len(val) < 3:
                        continue
                    for tm in matcher.find(val.upper()):
                        oid = db.get_or_create_object(tm.tag, tm.object_type)
                        coord = f"{ws.title}!{chr(65 + ci) if ci < 26 else 'C' + str(ci + 1)}{row_no}"
                        db.add_occurrence(oid, file_id, None, coord,
                                          context_snippet(val, tm.start, tm.end))
                        stats["tags"] += 1

            start_after = header_idx if header_idx is not None else -1
            for ri, row_vals in enumerate(cached):
                if ri > start_after or mode is None:
                    handle_row(list(row_vals) + [None] * MAX_SCAN_COLS, ri + 1)
            for ri, row_vals in enumerate(rows, start=len(cached)):
                handle_row(list(row_vals) + [None] * MAX_SCAN_COLS, ri + 1)

        db.set_file_flags(file_id, parsed=1,
                          doc_class="register" if stats["is_register"]
                          else ("model_export" if stats["is_model_export"] else None))
        return stats
    finally:
        wb.close()

"""
TRIZ Project Hub — drawing register generator.

Builds a styled Excel register straight from the index: latest revision per
document, full object index with occurrence counts, and the live issue list.
No manual transcription — the register is a *report on* the project, so it
can never drift from what's actually in the folder.
"""

from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2430")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)


def _sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY_FONT


def generate_register(db, out_path: str) -> str:
    wb = Workbook()

    ws = wb.active
    ws.title = "Drawing Register"
    docs = db.latest_documents()
    _sheet(ws,
           ["Document No", "Rev", "Title", "Type", "File", "Path", "Source"],
           [[d["doc_number"], d["revision"] or "", d["title"] or "",
             (d["doc_class"] or "").upper(), d["name"], d["path"], d["source"] or ""]
            for d in docs],
           [18, 6, 42, 12, 30, 55, 12])

    ws2 = wb.create_sheet("Object Index")
    objs = db.objects_list()
    _sheet(ws2,
           ["Tag", "Type", "In Model", "Occurrences"],
           [[o["tag"], o["type"], "Yes" if o["in_model"] else "", o["hits"]] for o in objs],
           [20, 14, 10, 12])

    ws3 = wb.create_sheet("Issues")
    issues = db.issues_list()
    _sheet(ws3,
           ["Severity", "Category", "Message", "File", "Tag"],
           [[i["severity"].upper(), i["category"], i["message"],
             i["name"] or "", i["tag"] or ""] for i in issues],
           [10, 18, 80, 28, 14])

    out = Path(out_path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


def default_register_name() -> str:
    return f"Drawing_Register_{time.strftime('%Y-%m-%d')}.xlsx"

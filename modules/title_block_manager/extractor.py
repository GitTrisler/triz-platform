from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    OPENPYXL_AVAILABLE = False

from core.autocad import (
    AutoCADSession,
    init_com_for_thread,
    uninit_com_for_thread,
    retry_busy_call,
    get_attr_retry,
    iter_com_collection,
)


DEFAULT_BLOCK_ALIASES = {
    "TITLEBLOCK": ["TITLEBLOCK", "TITLE BLOCK", "T2", "T3", "BORDER_24X36"],
}


class TitleBlockExtractor:
    def __init__(self, platform=None):
        self.platform = platform

    def write(self, message, level="INFO"):
        if self.platform:
            self.platform.output_write(message, level)
        else:
            print(f"[{level}] {message}")

    def extract_template(self, values: dict):
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl is not installed. Run: pip install openpyxl --break-system-packages"}

        drawing_folder = Path(values.get("drawing_folder", ""))
        output_folder = Path(values.get("output_folder", "")) if values.get("output_folder") else drawing_folder
        block_name = values.get("block_name", "TITLEBLOCK").strip() or "TITLEBLOCK"
        key_column = values.get("key_column", "CADFILE").strip() or "CADFILE"
        include_subfolders = bool(values.get("include_subfolders", True))

        if not drawing_folder.exists():
            return {"error": f"Drawing folder not found: {drawing_folder}"}

        output_folder.mkdir(parents=True, exist_ok=True)

        dwgs = sorted(
            drawing_folder.rglob("*.dwg")
            if include_subfolders
            else drawing_folder.glob("*.dwg")
        )

        if not dwgs:
            return {"error": "No DWG files found."}

        self.write("=" * 60, "INFO")
        self.write("Title Block Template Extraction started", "JOB")
        self.write(f"Drawing folder: {drawing_folder}", "INFO")
        self.write(f"DWGs found: {len(dwgs)}", "SUCCESS")
        self.write(f"Target block: {block_name}", "INFO")
        self.write("=" * 60, "INFO")

        session = AutoCADSession()
        results = []
        all_tags = []

        init_com_for_thread()

        try:
            session.start(visible=True)
            session.close_all()

            for index, dwg_path in enumerate(dwgs, start=1):
                self.write("", "INFO")
                self.write(f"[{index}/{len(dwgs)}] {dwg_path.name}", "JOB")

                doc = None

                try:
                    doc = session.open(dwg_path)

                    extracted = self._extract_from_document(
                        doc=doc,
                        dwg_path=dwg_path,
                        block_name=block_name,
                        key_column=key_column,
                    )

                    results.append(extracted)

                    for tag in extracted["attributes"].keys():
                        if tag not in all_tags:
                            all_tags.append(tag)

                    if extracted["attributes"]:
                        self.write(
                            f"  [OK] Extracted {len(extracted['attributes'])} attribute(s)",
                            "SUCCESS",
                        )
                    else:
                        self.write(
                            "  [WARN] No matching title block attributes found",
                            "WARNING",
                        )

                except Exception as e:
                    self.write(
                        f"  [ERROR] Failed extracting {dwg_path.name}: {e}",
                        "ERROR",
                    )

                    results.append({
                        key_column: dwg_path.stem,
                        "DWG_NAME": dwg_path.name,
                        "DWG_PATH": str(dwg_path),
                        "LAYOUT": "",
                        "BLOCK_NAME": block_name,
                        "attributes": {},
                        "error": str(e),
                    })

                finally:
                    if doc:
                        try:
                            session.close(doc, save=False)
                        except Exception:
                            pass

        finally:
            uninit_com_for_thread()

        template_path = output_folder / (
            f"TRIZ_TitleBlock_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        self._write_excel_template(
            template_path=template_path,
            results=results,
            all_tags=all_tags,
            key_column=key_column,
        )

        success_count = sum(1 for r in results if r.get("attributes"))
        failed_count = len(results) - success_count

        self.write("", "INFO")
        self.write("=" * 60, "INFO")
        self.write("Template extraction complete", "SUCCESS")
        self.write(f"Template saved: {template_path}", "SUCCESS")
        self.write(f"Drawings processed: {len(results)}", "INFO")
        self.write(f"Drawings with attributes: {success_count}", "SUCCESS")
        self.write(
            f"Drawings without attributes/errors: {failed_count}",
            "WARNING" if failed_count else "INFO",
        )

        return {
            "template_path": str(template_path),
            "drawings": len(results),
            "attributes": len(all_tags),
            "rows": len(results),
            "failed": failed_count,
        }

    def _extract_from_document(
        self,
        doc,
        dwg_path: Path,
        block_name: str,
        key_column: str,
    ):
        aliases = self._block_aliases(block_name)

        found_layout = ""
        found_block = ""
        attributes = {}

        layouts = get_attr_retry(doc, "Layouts")

        for layout in iter_com_collection(layouts):
            try:
                layout_name = get_attr_retry(layout, "Name")
                block = get_attr_retry(layout, "Block")
                block_count = get_attr_retry(block, "Count")

                for i in range(block_count):
                    entity = retry_busy_call(block.Item, i)

                    entity_name = get_attr_retry(entity, "EntityName")

                    if entity_name != "AcDbBlockReference":
                        continue

                    try:
                        effective_name = get_attr_retry(entity, "EffectiveName")
                    except Exception:
                        effective_name = get_attr_retry(entity, "Name")

                    if str(effective_name).strip().upper() not in aliases:
                        continue

                    has_attributes = get_attr_retry(entity, "HasAttributes")

                    if not has_attributes:
                        continue

                    found_layout = layout_name
                    found_block = str(effective_name)

                    attribute_refs = retry_busy_call(entity.GetAttributes)

                    for att in attribute_refs:
                        tag = str(get_attr_retry(att, "TagString")).strip().upper()
                        text_string = get_attr_retry(att, "TextString")
                        value = str(text_string) if text_string is not None else ""
                        attributes[tag] = value

                    return {
                        key_column: dwg_path.stem,
                        "DWG_NAME": dwg_path.name,
                        "DWG_PATH": str(dwg_path),
                        "LAYOUT": found_layout,
                        "BLOCK_NAME": found_block,
                        "attributes": attributes,
                        "error": "",
                    }

            except Exception:
                continue

        return {
            key_column: dwg_path.stem,
            "DWG_NAME": dwg_path.name,
            "DWG_PATH": str(dwg_path),
            "LAYOUT": found_layout,
            "BLOCK_NAME": found_block or block_name,
            "attributes": attributes,
            "error": "Title block not found",
        }

    def _write_excel_template(
        self,
        template_path: Path,
        results: list,
        all_tags: list,
        key_column: str,
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "Title Block Data"

        report = wb.create_sheet("Extraction Report")

        base_headers = [
            key_column,
            "DWG_NAME",
            "DWG_PATH",
            "LAYOUT",
            "BLOCK_NAME",
        ]

        headers = base_headers + all_tags + ["ERROR"]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        required_fill = PatternFill("solid", fgColor="2563EB")
        metadata_fill = PatternFill("solid", fgColor="374151")
        error_fill = PatternFill("solid", fgColor="7F1D1D")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for idx, cell in enumerate(ws[1], start=1):
            if idx == 1:
                cell.fill = required_fill
            elif cell.value in {"DWG_NAME", "DWG_PATH", "LAYOUT", "BLOCK_NAME"}:
                cell.fill = metadata_fill
            elif cell.value == "ERROR":
                cell.fill = error_fill

        for result in results:
            row = []

            for header in base_headers:
                row.append(result.get(header, ""))

            attributes = result.get("attributes", {})

            for tag in all_tags:
                row.append(attributes.get(tag, ""))

            row.append(result.get("error", ""))

            ws.append(row)

        report_headers = [
            "DWG_NAME",
            "DWG_PATH",
            "LAYOUT",
            "BLOCK_NAME",
            "ATTRIBUTES_FOUND",
            "STATUS",
            "ERROR",
        ]

        report.append(report_headers)

        for cell in report[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for result in results:
            attr_count = len(result.get("attributes", {}))
            error = result.get("error", "")
            status = "OK" if attr_count else "FAILED"

            report.append([
                result.get("DWG_NAME", ""),
                result.get("DWG_PATH", ""),
                result.get("LAYOUT", ""),
                result.get("BLOCK_NAME", ""),
                attr_count,
                status,
                error,
            ])

        self._autosize(ws)
        self._autosize(report)

        ws.freeze_panes = "A2"
        report.freeze_panes = "A2"

        wb.save(template_path)

    def _autosize(self, ws):
        for column_cells in ws.columns:
            max_length = 12

            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except Exception:
                    pass

            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max_length + 2,
                48,
            )

    def _block_aliases(self, block_name: str):
        key = block_name.strip().upper()

        aliases = {key}

        if key in DEFAULT_BLOCK_ALIASES:
            aliases.update(name.upper() for name in DEFAULT_BLOCK_ALIASES[key])

        return aliases
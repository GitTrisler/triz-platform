from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    load_workbook = None
    OPENPYXL_AVAILABLE = False


DEFAULT_SHEET_NAMES = [
    "Title Block Data",
    "TITLE BLOCK DATA",
    "Sheet1",
]

DEFAULT_KEY_COLUMNS = [
    "LINE NUMBER",
    "CADFILE",
    "TBDWGNO",
    "DrawingName",
    "DWG FILE",
    "DRAWING NUMBER",
]


def normalize_key(value) -> str:
    text = str(value or "").strip().lower()

    if text.endswith(".dwg"):
        text = text[:-4]

    return text


class TitleBlockScanner:
    def __init__(self, platform=None):
        self.platform = platform

    def write(self, message, level="INFO"):
        if self.platform:
            self.platform.output_write(message, level)
        else:
            print(f"[{level}] {message}")

    def scan(self, values: dict):
        if not OPENPYXL_AVAILABLE:
            return {
                "error": "openpyxl is not installed. Run: pip install openpyxl --break-system-packages"
            }

        excel_file = Path(values.get("excel_file", ""))
        drawing_folder = Path(values.get("drawing_folder", ""))
        worksheet_name = values.get("worksheet", "").strip()
        key_column = values.get("key_column", "").strip() or "CADFILE"
        include_subfolders = bool(values.get("include_subfolders", True))

        if not excel_file.exists():
            return {"error": f"Excel file not found: {excel_file}"}

        if not drawing_folder.exists():
            return {"error": f"Drawing folder not found: {drawing_folder}"}

        rows_by_key, worksheet_used = self._read_excel(
            excel_file=excel_file,
            worksheet_name=worksheet_name,
            key_column=key_column,
        )

        dwgs = self._scan_dwgs(
            drawing_folder=drawing_folder,
            include_subfolders=include_subfolders,
        )

        matched = []
        unmatched_dwgs = []

        for dwg in dwgs:
            key = normalize_key(dwg.stem)

            if key in rows_by_key:
                matched.append(dwg)
            else:
                unmatched_dwgs.append(dwg)

        matched_keys = {normalize_key(dwg.stem) for dwg in matched}
        unmatched_rows = [
            key for key in rows_by_key.keys()
            if key not in matched_keys
        ]

        result = {
            "rows": len(rows_by_key),
            "drawings": len(dwgs),
            "matched": len(matched),
            "unmatched_dwgs": len(unmatched_dwgs),
            "unmatched_rows": len(unmatched_rows),
            "worksheet": worksheet_used,
            "excel_file": str(excel_file),
            "drawing_folder": str(drawing_folder),
            "matches": matched,
        }

        self.write(f"Excel worksheet: {worksheet_used}", "INFO")
        self.write(f"Excel rows loaded: {result['rows']}", "SUCCESS")
        self.write(f"DWGs found: {result['drawings']}", "SUCCESS")
        self.write(f"Matched drawings: {result['matched']}", "SUCCESS")

        if unmatched_dwgs:
            self.write(f"DWGs with no Excel row: {len(unmatched_dwgs)}", "WARNING")

        if unmatched_rows:
            self.write(f"Excel rows with no DWG: {len(unmatched_rows)}", "WARNING")

        return result

    def _read_excel(self, excel_file: Path, worksheet_name: str, key_column: str):
        wb = load_workbook(str(excel_file), data_only=True, read_only=True)

        try:
            if worksheet_name:
                if worksheet_name not in wb.sheetnames:
                    raise RuntimeError(
                        f"Worksheet '{worksheet_name}' not found. "
                        f"Available sheets: {', '.join(wb.sheetnames)}"
                    )

                ws = wb[worksheet_name]

            else:
                ws = None

                for name in DEFAULT_SHEET_NAMES:
                    if name in wb.sheetnames:
                        ws = wb[name]
                        break

                if ws is None:
                    ws = wb.active

            header_info = self._find_header_row(ws, key_column)

            if not header_info:
                raise RuntimeError(
                    f"Could not find key column '{key_column}' in workbook."
                )

            headers, header_row_index, key_index = header_info

            rows_by_key = {}

            for row in ws.iter_rows(
                min_row=header_row_index + 1,
                values_only=True,
            ):
                key_value = row[key_index] if key_index < len(row) else None
                key = normalize_key(key_value)

                if not key:
                    continue

                row_data = {}

                for i, header in enumerate(headers):
                    if not header:
                        continue

                    value = row[i] if i < len(row) else None
                    row_data[str(header).strip().upper()] = value

                rows_by_key[key] = row_data

            return rows_by_key, ws.title

        finally:
            wb.close()

    def _find_header_row(self, ws, key_column: str):
        key_upper = key_column.strip().upper()
        fallback_keys = [k.upper() for k in DEFAULT_KEY_COLUMNS]

        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            headers = [
                str(cell).strip() if cell is not None else ""
                for cell in row
            ]

            upper_headers = [h.upper() for h in headers]

            if key_upper in upper_headers:
                return headers, row_index, upper_headers.index(key_upper)

            for fallback in fallback_keys:
                if fallback in upper_headers:
                    return headers, row_index, upper_headers.index(fallback)

        return None

    def _scan_dwgs(self, drawing_folder: Path, include_subfolders: bool):
        if include_subfolders:
            return sorted(drawing_folder.rglob("*.dwg"))

        return sorted(drawing_folder.glob("*.dwg"))
